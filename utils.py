"""
Scipy version > 0.18 is needed, due to 'mode' option from scipy.misc.imread function
"""

import os
import numpy as np
import tensorflow as tf
import cv2
import shutil
import pandas as pd
from openpyxl import load_workbook

FLAGS = tf.app.flags.FLAGS
def gaussian_noise( input, std=0.05):
    noise = tf.random_normal(shape=tf.shape(input), mean=0.0, stddev=std, dtype=tf.float32)
    return input + noise

def form_results(results_path='./Results', model_type=None, dataset=None):
    """
    Forms folders for each run to store the tensorboard files, saved models and the log files.
    :return: three string pointing to tensorboard, saved models and log paths respectively.
    """
    if not os.path.exists(results_path):
        os.mkdir(results_path)
    folder_name = "/{0}_{1}_model".format(model_type, dataset)
    tensorboard_path = results_path + folder_name + '/Tensorboard'
    log_path = results_path + folder_name + '/log'
    if os.path.exists(results_path + folder_name):
        shutil.rmtree(results_path + folder_name)
    if not os.path.exists(results_path + folder_name):
        os.mkdir(results_path + folder_name)
        os.mkdir(tensorboard_path)
        os.mkdir(log_path)
    return tensorboard_path,  log_path

def RGB2YCbCr(RGB_image):
    ## RGB_image [-1, 1]
    test_num1 = 16.0 / 255.0
    test_num2 = 128.0 / 255.0
    R = RGB_image[:, :, :, 0:1]
    G = RGB_image[:, :, :, 1:2]
    B = RGB_image[:, :, :, 2:3]
    Y = 0.257 * R + 0.564 * G + 0.098 * B + test_num1
    Cb = - 0.148 * R - 0.291 * G + 0.439 * B + test_num2
    Cr = 0.439 * R - 0.368 * G - 0.071 * B + test_num2
    return Y, Cb, Cr

def RGB2Gray(RGB_image):
    RGB_image = RGB_image * 255.0


def YCbCr2RGB(Y, Cb, Cr, mode=1):
    ## Y, Cb, Cr :[-1, 1]
    test_num1 = 16.0 / 255.0
    test_num2 = 128.0 / 255.0
    R = 1.164 * (Y - test_num1) + 1.596 * (Cr - test_num2)
    G = 1.164 * (Y - test_num1) - 0.392 * (Cb - test_num2) - 0.813 * (Cr - test_num2)
    B = 1.164 * (Y - test_num1) + 2.017 * (Cb - test_num2)
    RGB_image = tf.concat([R, G, B], axis=-1)
    BGR_image = tf.concat([B, G, R], axis=-1)
    if mode == 1:
        return RGB_image
    else:
        return BGR_image

def illumination_mechanism(day_probability, night_probability, scheme_num=1):
    if scheme_num == 1:
        vi_w = day_probability / tf.add(day_probability, night_probability)
        ir_w = night_probability / tf.add(day_probability, night_probability)
    elif scheme_num == 2:
        vi_w = tf.exp(day_probability) / tf.add(tf.exp(day_probability), tf.exp(night_probability))
        ir_w = tf.exp(night_probability) / tf.add(tf.exp(day_probability), tf.exp(night_probability))
    elif scheme_num == 3:
        day_probability = tf.log(day_probability)
        night_probability = tf.log(night_probability)
        vi_w = day_probability / tf.add(day_probability, night_probability)
        ir_w = night_probability / tf.add(day_probability, night_probability)
    elif scheme_num == 4:
        min_labels = 0.1 * tf.ones_like(day_probability)
        max_labels = 0.9 * tf.ones_like(day_probability)

        min_labels_2 = 0.4 * tf.ones_like(day_probability)
        max_labels_2 = 0.6 * tf.ones_like(day_probability)
        vi_w = tf.where(day_probability - night_probability>0, max_labels, min_labels)
        ir_w = tf.where(day_probability - night_probability>0, min_labels, max_labels)
    return vi_w, ir_w
def gradient(input):
    filter1 = tf.reshape(tf.constant([[-1., 0., 1.], [-2., 0., 2.], [-1., 0., 1.]]), [3, 3, 1, 1])
    filter2 = tf.reshape(tf.constant([[-1., -2., -1.], [0., 0., 0.], [1., 2., 1.]]), [3, 3, 1, 1])
    Gradient1 = tf.nn.conv2d(input, filter1, strides=[1, 1, 1, 1], padding='SAME')
    Gradient2 = tf.nn.conv2d(input, filter2, strides=[1, 1, 1, 1], padding='SAME')
    Gradient = tf.abs(Gradient1) + tf.abs(Gradient2)
    return Gradient


def weights_spectral_norm(weights, u=None, iteration=1, update_collection=None, reuse=False, name='weights_SN'):
    with tf.variable_scope(name) as scope:
        if reuse:
            scope.reuse_variables()

        w_shape = weights.get_shape().as_list()
        w_mat = tf.reshape(weights, [-1, w_shape[-1]])
        if u is None:
            u = tf.get_variable('u', shape=[1, w_shape[-1]], initializer=tf.truncated_normal_initializer(),
                                trainable=False)

        def power_iteration(u, ite):
            v_ = tf.matmul(u, tf.transpose(w_mat))
            v_hat = l2_norm(v_)
            u_ = tf.matmul(v_hat, w_mat)
            u_hat = l2_norm(u_)
            return u_hat, v_hat, ite + 1

        u_hat, v_hat, _ = power_iteration(u, iteration)

        sigma = tf.matmul(tf.matmul(v_hat, w_mat), tf.transpose(u_hat))

        w_mat = w_mat / sigma

        if update_collection is None:
            with tf.control_dependencies([u.assign(u_hat)]):
                w_norm = tf.reshape(w_mat, w_shape)
        else:
            if not (update_collection == 'NO_OPS'):
                print(update_collection)
                tf.add_to_collection(update_collection, u.assign(u_hat))

            w_norm = tf.reshape(w_mat, w_shape)
        return w_norm


def lrelu(x, leak=0.2):
    return tf.maximum(x, leak * x)

def l2_norm(input_x, epsilon=1e-12):
    input_x_norm = input_x / (tf.reduce_sum(input_x ** 2) ** 0.5 + epsilon)
    return input_x_norm

def check_folder(dir):
    if not os.path.exists(dir):
        os.makedirs(dir)

def load_test_data(image_path, mode=1):

    if mode == 1:
        print('image_path: ', image_path)
        img = cv2.imread(image_path, 0)
        print('image shape: ', img.shape)
        img = np.expand_dims(img, axis=0)
        img = np.expand_dims(img, axis=-1)
        img = preprocessing(img)
    else:
        print('image_path: ', image_path)
        img = cv2.imread(image_path, 1)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB).astype(np.float32)
        print('image shape: ', img.shape)
        img = np.expand_dims(img, axis=0)
        img = preprocessing(img)
    return img

def preprocessing(x):
    x = x / 255.0 # -1 ~ 1
    return x

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def writexls(save_name, method_name, time_list, sheet_name, i):
    df = pd.DataFrame({method_name: time_list})
    append_df_to_excel(save_name, df, sheet_name=sheet_name, index=False, startrow=0, startcol=i)